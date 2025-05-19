
import openpyxl 
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles.borders import Border, Side

import sys
import importlib


import http.cookiejar

import tkinter as tk


cookielib = http.cookiejar

class FG:
    m_FG_Code = ''                      #    599699
    m_FG_Desc = ''                      #    
    m_FG_Category = ''
    m_Ingredient = list()
    m_Line_A =''
    m_Line_B =''
    m_Line_C =''
    m_Line_D =''
    m_Line_E =''

    def FG():
        m_FG_Code = ''                  #   SOAM00000099
        m_FG_Desc = 0                   #   Internal Order Number        
        m_FG_Category = ''
        m_Ingredients = list()
        m_Line_A =''
        m_Line_B =''
        m_Line_C =''
        m_Line_D =''
        m_Line_E =''

class Ingredient:
    m_Name = ''                     #   Item Name
    m_Description = ''              #   Pending/Delivered
    m_ProductCategory=''            #   Product category
    m_Quantity = 0                  #   Ordered quantity
    m_Committed = 0                 #   Committed number
    m_BackOrdered = 0               #   Backordered number
    m_InventoryLocation = ''        #   Inventory Location
    m_Picked = 0                    #   Picked quantity
    m_Packed = 0                    #   Packed quantity
    m_Fulfilled = 0                 #   Fulfilled quantity

    def reset():
        m_Name = ''                     #   Item Name
        m_Description = ''              #   Pending/Delivered
        m_ProductCatetory = ''          #   Product category
        m_Quantity = 0                  #   Ordered quantity
        m_Committed = 0                 #   Committed number
        m_BackOrdered = 0               #   Backordered number
        m_InventoryLocation = ''        #   Inventory Location

g_FG_List            = []

g_Iterater_Start        = 0
g_Iterater_Stop         = 9999
g_Iterater_Index        = 0
g_Use_Local_File        = 1


g_Fill_Cell_Red     = PatternFill(patternType='solid', fgColor='F8CBAD')
g_Fill_Cell_Green_L = PatternFill(patternType='solid', fgColor='C6E0B4')
g_Fill_Cell_Green_D = PatternFill(patternType='solid', fgColor='A9D08E')
g_Fill_Cell_Blue_L  = PatternFill(patternType='solid', fgColor='BDD7EE')
g_Fill_Cell_Blue_D  = PatternFill(patternType='solid', fgColor='8EA9DB')
g_Fill_Cell_Yellow  = PatternFill(patternType='solid', fgColor='FFFF00')
g_Fill_Cell_Grey    = PatternFill(patternType='solid', fgColor='D9D9D9')
g_Fill_Cell_Purpal  = PatternFill(patternType='solid', fgColor='7030A0')

g_Thin_Border = Border(left     =Side(style='thin'), 
                       right    =Side(style='thin'), 
                       top      =Side(style='thin'), 
                       bottom   =Side(style='thin'))

g_RootDir = "f://PMF//"
g_Recipe_Folder = "Recipes//"
g_Username = ""
g_Password = ""


def aa_cell(wb,r,c,v,alignment='left'):
    wb.cell(r,c,v)
    wb.cell(r,c).border = g_Thin_Border
    wb.cell(r,c).alignment = Alignment(wrapText=True, horizontal=alignment)    

def aa_upper(val):
    if(val==None):
        return ""
    else:
        return str(val).strip().upper()

def aa_value(val):
    if(val==None):
        return ""
    else:
        return str(val).strip().upper()



def read_excel_xls_finish_goods_list(path):

    i_workbook      = openpyxl.load_workbook(path, data_only=True)      # 打开工作簿
    i_worksheet     = i_workbook.worksheets[0]
    i_rows          = i_worksheet.rows

    i_col_I         = -1
    i_col_Code      = -1
    i_col_Desc      = -1
    i_col_Category  = -1
    i_col_Allergen  = -1
    i_col_LineA     = -1
    i_col_LineB     = -1
    i_col_LineC     = -1
    i_col_LineD     = -1
    i_col_LineE     = -1

    i_str_Index         = 'INDEX'
    i_str_Code          = 'CODE'
    i_str_Desc          = 'DESC'
    i_str_Category      = 'CATEGORY'
    i_str_Allergen      = 'ALLERGEN'
    i_str_LineA         = 'LINEA'
    i_str_LineB         = 'LINEB'
    i_str_LineC         = 'LINEC'
    i_str_LineD         = 'LINED'
    i_str_LineE         = 'LINEE'

    i_row_index         = 0
    i_col_index         = 0    

    for i_row in i_rows:

        i_fg = FG()

        i_col_index = 0

        for i_cell in i_row:
            if(i_row_index==0):
                if aa_upper(i_cell.value)==i_str_Index      : i_col_I           = i_col_index
                if aa_upper(i_cell.value)==i_str_Code       : i_col_Code        = i_col_index
                if aa_upper(i_cell.value)==i_str_Desc       : i_col_Desc        = i_col_index
                if aa_upper(i_cell.value)==i_str_Category   : i_col_Category    = i_col_index
                if aa_upper(i_cell.value)==i_str_Allergen   : i_col_Allergen    = i_col_index
                if aa_upper(i_cell.value)==i_str_LineA      : i_col_LineA       = i_col_index
                if aa_upper(i_cell.value)==i_str_LineB      : i_col_LineB       = i_col_index
                if aa_upper(i_cell.value)==i_str_LineC      : i_col_LineC       = i_col_index
                if aa_upper(i_cell.value)==i_str_LineD      : i_col_LineD       = i_col_index
                if aa_upper(i_cell.value)==i_str_LineE      : i_col_LineE       = i_col_index
            else:
                if(i_col_index == i_col_Code):
                    i_fg.m_FG_Code = aa_upper(i_cell.value)

                if(i_col_index == i_col_Desc):
                    i_fg.m_FG_Desc = aa_upper(i_cell.value)

                if(i_col_index == i_col_Category):
                    i_fg.m_FG_Category = aa_upper(i_cell.value)

                if(i_col_index == i_col_Allergen):
                    i_fg.m_Ingredient = aa_upper(i_cell.value)

                if(i_col_index == i_col_LineA):
                    i_fg.m_Line_A = aa_upper(i_cell.value)

                if(i_col_index == i_col_LineB):
                    i_fg.m_Line_B = aa_upper(i_cell.value)

                if(i_col_index == i_col_LineC):
                    i_fg.m_Line_C = aa_upper(i_cell.value)

                if(i_col_index == i_col_LineD):
                    i_fg.m_Line_D = aa_upper(i_cell.value)

                if(i_col_index == i_col_LineE):
                    i_fg.m_Line_E = aa_upper(i_cell.value)

            i_col_index = i_col_index + 1

        if(i_fg.m_FG_Code != ''):
            i_fg.m_Ingredients = list()
            g_FG_List.append(i_fg)

        i_row_index = i_row_index + 1

def ana_page(delivery_list_file_path):
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    import time

    global g_RootDir
    global g_Recipe_Folder

    # 创建浏览器实例
    driver = webdriver.Chrome()  # 需要先安装 Chrome 浏览器驱动

    # 导航到登录页面
    login_url = "https://pmf.makeithappen.com/"
    driver.get(login_url)

    # 找到用户名和密码输入字段的元素，并输入对应的值
    username_field = driver.find_element(By.XPATH, "//input[@placeholder='User ID / Email / Mobile No.']")
    password_field = driver.find_element(By.XPATH, "//input[@placeholder='Password']")
    username_field.send_keys(g_Username)
    password_field.send_keys(g_Password)

    # 提交登录表单
    login_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Login')]")
    login_button.click()

    time.sleep(3)

    # 可以继续进行其他操作，例如访问其他页面
    for i_fg in g_FG_List:
        driver.get("https://pmf.makeithappen.com/procurement/1/bom/"+i_fg.m_FG_Code)
        time.sleep(1)
        page_content = driver.page_source

        if(i_fg.m_FG_Category=="WIP"):
            i_local_file_path = g_RootDir + g_Recipe_Folder + "WIP/" + i_fg.m_FG_Code + ".HTML"
        else:
            i_local_file_path = g_RootDir + g_Recipe_Folder + i_fg.m_FG_Code + ".HTML"


        with open(i_local_file_path, 'w', encoding='utf-8') as i_local_f:
                    i_local_f.write(page_content)
                    i_local_f.close()

    # soup = BeautifulSoup(page_content, 'html.parser')

    # # 示例操作：找到页面中的标题
    # stock_label = soup.find(text='Stock')
    # stock_input = stock_label.find_next_sibling('input')
    # stock_value = stock_input['value']

    # 关闭浏览器实例
    driver.quit()


def _write_excel_xls_Recipe_list(so, Recipe_list_file_path):

    i_workbook = openpyxl.load_workbook(Recipe_list_file_path)    # 打开工作簿    
    i_worksheet = i_workbook.worksheets[0]    
    i_row_count = i_worksheet.max_row + 1

    ###################################
    #
    #    Index
    #
    ###################################
    # aa_cell(i_worksheet, i_row_count, g_Col_Output_Index,i_row_count-1,'center')
    aa_cell(i_worksheet, i_row_count, g_Col_Output_Index,'=ROW()-1','center')

    ###################################
    #
    #    SAVE
    #
    ###################################
    i_workbook.save(Recipe_list_file_path)  # 保存工作簿

class ConfigWindow:
    def __init__(self, master):
        self.master = master
        master.title("Configuration")

        # Create widgets
        self.init_file_label = tk.Label(master, text="init file:")        
        self.init_file_text = tk.StringVar()
        self.init_file_text.set("c:\\pcg\\init.xlsx")
        self.init_file_entry = tk.Entry(master, textvariable=self.init_file_text, width=100)

        # self.age_label = tk.Label(master, text="Age:")
        # self.age_entry = tk.Entry(master)
        self.submit_button = tk.Button(master, text="Submit", command=self.submit)

        # Lay out widgets
        self.init_file_label.grid(row=0, column=0)
        self.init_file_entry.grid(row=0, column=1)
        # self.age_label.grid(row=1, column=0)
        # self.age_entry.grid(row=1, column=1)
        self.submit_button.grid(row=2, column=0, columnspan=2)

    def submit(self):
        i_str_finish_goods_path         = '00 fg-list.xlsx'
        i_str_Recipe_list_file_path     = '00 Recipe-list.xlsx'
    
        # init("c:\\pcg\\init.xlsx")
        init(self.init_file_text.get())

        i_str_finish_goods_path         = g_RootDir + i_str_finish_goods_path
        i_str_Recipe_list_file_path     = g_RootDir + i_str_Recipe_list_file_path

        read_excel_xls_finish_goods_list(i_str_finish_goods_path)    
        ana_page(i_str_Recipe_list_file_path)

def init(path):
    global g_RootDir
    global g_Recipe_Folder
    global g_Username
    global g_Password

    i_workbook      = openpyxl.load_workbook(path, data_only=True)      # 打开工作簿

    i_worksheet     = i_workbook.worksheets[0]
    i_rows          = i_worksheet.rows
    
    i_col_index = 0

    g_RootDir       = i_worksheet['B1'].value
    g_Recipe_Folder = i_worksheet['B2'].value
    g_Username      = i_worksheet["B3"].value
    g_Password      = i_worksheet["B4"].value

def main():
    importlib.reload(sys)

    i_str_finish_goods_list         = '00 fg-list.xlsx'
    i_str_Recipe_list               = '00 Recipe-list.xlsx'

    i_str_finish_goods_path         = g_RootDir + i_str_finish_goods_list
    i_str_Recipe_list_file_path     = g_RootDir + i_str_Recipe_list

    root = tk.Tk()
    config_window = ConfigWindow(root)
    root.mainloop()

if __name__ == '__main__':
    main()